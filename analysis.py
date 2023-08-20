# output.xlsx 파일을 읽는 python 프로그램


from openpyxl import load_workbook

wb = load_workbook('output.xlsx')
ws = wb.active

# ws의 컬럼 수와 행 수를 출력
print(ws.max_column, ws.max_row)

# ws의 데이터를 dataframe으로 변환
import pandas as pd
df = pd.DataFrame(ws.values)

# df의 컬럼 이름을 변경
df.columns = ['번호', '제목', '첨부파일', '등록일', '조회수', '본문']

# 첨부파일 컬럼 삭제
df.drop('첨부파일', axis=1, inplace=True)

# 등록일 컬럼이 OOOO.OO.OO 형식이므로 OOOO-OO-OO 형식으로 변경
df['등록일'] = df['등록일'].apply(lambda x: x.replace('.', '-'))

# 등록일이 2021년 7월 12일 이전인 데이터는 삭제
df = df[df['등록일'] > '2021-07-12']

# df에 소관부서 컬럼을 추가하는데, 
# 본문 컬럼에 '산업안전보건정책과', '산재보상정책과', '산업안전기준과', '산업보건기준과', '직업건강증진팀',
# '안전보건감독기획과' ,'산재예방지원과', '건설산재예방정책과' ,'중대산업재해감독과' ,'화학사고예방과' 가 포함되어 있으면 소관부서 명에 해당 문자열을 추가
# 없으면 '기타'를 추가하고, 둘 이상이 있는 경우에는 둘 다 추가
df['소관부서'] = df['본문'].apply(lambda x: '산업안전보건정책과' if '산업안전보건정책과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '산재보상정책과' if '산재보상정책과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '산업안전기준과' if '산업안전기준과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '산업보건기준과' if '산업보건기준과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '직업건강증진팀' if '직업건강증진팀' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '안전보건감독기획과' if '안전보건감독기획과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '산재예방지원과' if '산재예방지원과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '건설산재예방정책과' if '건설산재예방정책과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '중대산업재해감독과' if '중대산업재해감독과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '화학사고예방과' if '화학사고예방과' in x else '')
df['소관부서'] += df['본문'].apply(lambda x: '기타' if '산업안전보건정책과' not in x and '산재보상정책과' not in x and '산업안전기준과' not in x and '산업보건기준과' not in x and '직업건강증진팀' not in x and '안전보건감독기획과' not in x and '산재예방지원과' not in x and '건설산재예방정책과' not in x and '중대산업재해감독과' not in x and '화학사고예방과' not in x else '')

# 등록일 컬럼에서 등록년도, 등록월 컬럼 생성
df['등록년도'] = df['등록일'].apply(lambda x: x.split('-')[0])
df['등록월'] = df['등록일'].apply(lambda x: x.split('-')[1])

# df의 컬럼 순서를 변경
df = df[['번호', '제목', '본문', '등록일', '등록년도', '등록월', '조회수', '소관부서']]
print(df.head())

# df에 '참고' 컬럼을 생성하고, 제목에 (참고)가 포함되어 있으면 '참고'를 추가하고 아니면 '보도'를 추가
df['참고'] = df['제목'].apply(lambda x: '참고' if '(참고)' in x else '보도')

# output2.xlsx 파일로 저장
df.to_excel('output2.xlsx', index=True)


















wb.close()

# Path: analysis.py
# 위의 코드를 실행하면 다음과 같은 결과가 출력됩니다.
